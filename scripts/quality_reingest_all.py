from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from sqlalchemy import func, select, text
from sqlalchemy.orm.attributes import flag_modified

from backend.app import models
from backend.app.config import get_settings
from backend.app.db import SessionLocal, create_all
from backend.app.services.asset_ingestor import create_asset_from_ingested_payload
from backend.app.services.asset_ingestor import prepare_asset_payload
from backend.app.services.geocode import geocode_address, google_maps_link
from backend.app.services.google_sheets_sync import _normalize_columns as gs_normalize_columns
from backend.app.services.google_sheets_sync import _row_payload as gs_row_payload
from backend.app.services.google_sheets_sync import _values_for_tab
from backend.app.services.ingestion import clean_cell, dedupe_fingerprint, is_blank_row, normalize_text, row_uid
from scripts.import_pearl_spytech_notes import (
    attachment_index,
    older_deal_payload,
    parse_workbook,
    property_payload,
    source_uid as pearl_source_uid,
)


SOURCE_WORKBOOKS = [
    ROOT / "data/uploads/Land Plot Manager.xlsx",
    ROOT / "data/uploads/Pearl Spytech Notes Tracker.xlsx",
]


def clean_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text_value = str(value)
    text_value = text_value.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text_value)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def parse_money(value: Any, *, raw: dict[str, Any] | None = None, field_name: str = "") -> float | None:
    amount = clean_number(value)
    if amount is None:
        return None
    text_value = normalize_text(value)
    if any(token in text_value for token in ["cr", "crore", "crores"]):
        return amount * 10_000_000
    if any(token in text_value for token in ["lac", "lakh", "lakhs"]):
        return amount * 100_000
    if "₹" in str(value) and amount < 1_000 and field_name.lower() in {"price", "asking_price"}:
        return amount * 10_000_000

    raw = raw or {}
    unit_price = clean_number(raw.get("Price per unit"))
    area = clean_number(raw.get("Size (for calculation)") or raw.get("%") or raw.get("Acreage"))
    if unit_price and area and amount < 1_000 and field_name.lower() in {"price", "asking_price"}:
        calculated = unit_price * area
        if calculated > 100_000:
            return calculated
    return amount


def normalize_status(value: Any) -> str:
    text_value = normalize_text(value)
    if not text_value:
        return "lead"
    if text_value in {"yes", "true", "active", "working", "working actively"}:
        return "active"
    if text_value in {"no", "false", "0"}:
        return "lead"
    if "sold" in text_value:
        return "sold"
    if "hold" in text_value:
        return "on_hold"
    return text_value.replace(" ", "_")


def normalize_payload(payload: dict[str, Any], *, source: str) -> dict[str, Any]:
    raw = payload.get("raw_source") if isinstance(payload.get("raw_source"), dict) else {}
    payload = {key: value for key, value in payload.items() if value not in (None, "", [])}
    title = payload.get("title") or raw.get("LOCATION") or raw.get("Property / Deal Name") or raw.get("Note Title")
    locality = payload.get("locality") or raw.get("LOCATION") or raw.get("Location / Area")
    district = payload.get("district") or raw.get("District")
    if not district and any("jaipur" in normalize_text(part) for part in [title, locality, payload.get("address"), raw]):
        district = "Jaipur"
    address = payload.get("address") or locality or title

    raw_price = payload.get("asking_price") or raw.get("PRICE") or payload.get("price_terms") or raw.get("Price / Commercial Terms")
    expected_price = payload.get("expected_price")
    payload["title"] = str(title or "Imported property").strip()
    payload["locality"] = locality
    payload["address"] = address
    payload["district"] = district
    payload["state"] = payload.get("state") or "Rajasthan"
    payload["status"] = normalize_status(raw.get("Working Actively?") or payload.get("status"))
    payload["asking_price"] = parse_money(raw_price, raw=raw, field_name="asking_price")
    payload["expected_price"] = parse_money(expected_price, raw=raw, field_name="expected_price")
    payload["workability_rating"] = int(clean_number(raw.get("Rating (Subject to interest)") or payload.get("workability_rating")) or 0) or None
    if raw.get("Hangup") and "Hangup:" not in str(payload.get("bottleneck_notes") or ""):
        payload["bottleneck_notes"] = "\n\n".join(
            str(part)
            for part in [payload.get("bottleneck_notes"), f"Hangup: {raw.get('Hangup')}"]
            if part
        )
    payload["source"] = source
    payload["approval_status"] = "approved"
    payload["raw_source"] = raw
    return {key: value for key, value in payload.items() if value not in (None, "", [])}


def google_sheet_records() -> list[dict[str, Any]]:
    settings = get_settings()
    records: list[dict[str, Any]] = []
    if not settings.google_sheet_id:
        return records
    for tab_name in [tab.strip() for tab in settings.google_sheet_tabs.split(",") if tab.strip()]:
        values = _values_for_tab(settings.google_sheet_id, tab_name)
        if not values:
            continue
        columns = [str(value).strip() if value not in (None, "") else f"Column {index + 1}" for index, value in enumerate(values[0])]
        mapping = gs_normalize_columns(columns)
        for row_number, row in enumerate(values[1:], start=2):
            raw = {
                columns[index] if index < len(columns) else f"Column {index + 1}": clean_cell(value)
                for index, value in enumerate(row)
            }
            if is_blank_row(raw):
                continue
            payload = gs_row_payload(raw, mapping, "google_sheets")
            records.append(
                {
                    "source": "google_sheets",
                    "source_uid": row_uid("google_sheets", f"{settings.google_sheet_id}:{tab_name}", row_number, raw),
                    "source_name": f"Google Sheet: {tab_name}",
                    "payload": normalize_payload(payload, source="google_sheets"),
                }
            )
    return records


def generic_excel_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []
    columns = [str(value).strip() if value not in (None, "") else f"Column {index + 1}" for index, value in enumerate(header_row)]
    mapping = gs_normalize_columns(columns)
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        raw = {
            columns[index] if index < len(columns) else f"Column {index + 1}": clean_cell(value)
            for index, value in enumerate(row)
        }
        if is_blank_row(raw):
            continue
        payload = gs_row_payload(raw, mapping, "excel")
        records.append(
            {
                "source": "excel",
                "source_uid": row_uid("excel", path.name, row_number, raw),
                "source_name": path.name,
                "payload": normalize_payload(payload, source="excel"),
            }
        )
    return records


def pearl_workbook_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    sheets = parse_workbook(path)
    serial_docs, url_docs = attachment_index(sheets.get("Attachments Index", []))
    records: list[dict[str, Any]] = []
    for index, row in enumerate(sheets.get("Property Notes", []), start=1):
        payload = property_payload(row, serial_docs, url_docs)
        records.append(
            {
                "source": "excel_pearl_spytech_notes",
                "source_uid": pearl_source_uid("excel_pearl_spytech_notes", row, f"property:{index}"),
                "source_name": path.name,
                "payload": normalize_payload(payload, source="excel_pearl_spytech_notes"),
            }
        )
    for index, row in enumerate(sheets.get("Investible Deals Older", []), start=1):
        payload = older_deal_payload(row)
        digest = json.dumps(row, sort_keys=True, ensure_ascii=False)
        records.append(
            {
                "source": "excel_investible_deals_older",
                "source_uid": f"excel_investible_deals_older:{dedupe_fingerprint({'title': digest})}",
                "source_name": path.name,
                "payload": normalize_payload(payload, source="excel_investible_deals_older"),
            }
        )
    return records


def backup_database(db) -> Path:
    import pandas as pd

    def safe_cell(value: Any) -> Any:
        if isinstance(value, datetime):
            return value.replace(tzinfo=None).isoformat(sep=" ")
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return value

    export_dir = ROOT / "data/exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / f"pre_quality_reingest_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    table_names = [
        "assets",
        "contacts",
        "asset_contacts",
        "owners",
        "brokers",
        "asset_documents",
        "asset_locations",
        "asset_updates",
        "approval_queue",
        "ingestion_logs",
        "notion_sync_logs",
    ]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for table_name in table_names:
            rows = db.execute(text(f"SELECT * FROM {table_name}")).mappings().all()
            pd.DataFrame(
                [{key: safe_cell(value) for key, value in dict(row).items()} for row in rows]
            ).to_excel(writer, sheet_name=table_name[:31], index=False)
    return path


def clear_ingested_data(db) -> None:
    for table in [
        "asset_match_suggestions",
        "deals",
        "asset_tags",
        "asset_updates",
        "asset_locations",
        "asset_documents",
        "asset_contacts",
        "assets",
        "contacts",
        "owners",
        "brokers",
        "approval_queue",
        "ingestion_logs",
        "notion_sync_logs",
    ]:
        db.execute(text(f"DELETE FROM {table}"))
    db.commit()


def local_location_score(asset: models.Asset) -> tuple[float, str]:
    text_value = normalize_text(" ".join(str(part) for part in [asset.address, asset.locality, asset.district] if part))
    if not text_value:
        return 3.0, "Minimal location data."
    premium = {
        "c-scheme": 9.5,
        "c scheme": 9.5,
        "civil lines": 9.2,
        "bani park": 8.8,
        "vaishali nagar": 8.7,
        "malviya nagar": 8.6,
        "jln marg": 9.0,
        "tonk road": 8.2,
        "mansarovar": 8.0,
        "raja park": 8.3,
        "bapu nagar": 8.5,
    }
    for zone, score in premium.items():
        if zone in text_value:
            return score, f"Premium zone match: {zone.title()}."
    score = 3.0
    if asset.district:
        score += 1.5
    if asset.locality:
        score += 1.5
    if asset.address and len(asset.address) > 15:
        score += 1.0
    if asset.latitude and asset.longitude:
        score += 1.0
    return round(min(score, 8.0), 1), "Rule-based score from location detail and coordinates."


def enrich_maps(db) -> dict[str, int]:
    assets = db.scalars(select(models.Asset).order_by(models.Asset.id)).all()
    cache: dict[str, tuple[float | None, float | None]] = {}
    geocoded = scored = linked = 0
    for asset in assets:
        query = ", ".join(
            str(part)
            for part in [asset.address, asset.locality, asset.tehsil, asset.district, asset.state or "Rajasthan", "India"]
            if part
        )
        if (asset.latitude is None or asset.longitude is None) and query:
            if query not in cache:
                cache[query] = geocode_address(query)
            lat, lon = cache[query]
            if lat is not None and lon is not None:
                asset.latitude = lat
                asset.longitude = lon
                geocoded += 1
        if asset.latitude is not None and asset.longitude is not None:
            asset.google_maps_link = google_maps_link(asset.latitude, asset.longitude)
            linked += 1
        raw_source = asset.raw_source if isinstance(asset.raw_source, dict) else {}
        ingestion_meta = raw_source.get("_ingestion") if isinstance(raw_source.get("_ingestion"), dict) else {}
        ingestion_meta["skip_geocode"] = False
        raw_source["_ingestion"] = ingestion_meta
        score, reason = local_location_score(asset)
        raw_source["location_score"] = score
        raw_source["location_score_reason"] = reason
        asset.raw_source = dict(raw_source)
        flag_modified(asset, "raw_source")
        scored += 1
        primary = next((location for location in asset.locations if location.label == "Primary"), None)
        if primary:
            primary.address = asset.address
            primary.latitude = asset.latitude
            primary.longitude = asset.longitude
            primary.google_maps_link = asset.google_maps_link
        else:
            db.add(
                models.AssetLocation(
                    asset_id=asset.id,
                    label="Primary",
                    address=asset.address,
                    latitude=asset.latitude,
                    longitude=asset.longitude,
                    google_maps_link=asset.google_maps_link,
                )
            )
        if scored % 50 == 0:
            db.commit()
            print(f"map_enrichment_progress={scored}/{len(assets)}")
    db.commit()
    return {"assets_seen": len(assets), "geocoded_missing": geocoded, "maps_linked": linked, "scored": scored}


def publish_records(db, records: list[dict[str, Any]]) -> dict[str, Any]:
    seen_fingerprints: set[str] = set()
    seen_source_uids: set[tuple[str, str]] = set()
    skipped = failed = 0
    failures: list[dict[str, str]] = []
    source_counts: Counter[str] = Counter()
    next_code = (db.scalar(select(func.count(models.Asset.id))) or 0) + 1
    asset_columns = {
        column.name
        for column in models.Asset.__table__.columns
        if column.name not in {"id", "created_at", "updated_at"}
    }
    asset_rows: list[dict[str, Any]] = []
    document_rows_by_code: list[tuple[str, dict[str, Any]]] = []
    update_rows_by_code: list[tuple[str, str]] = []
    for record in records:
        payload = dict(record["payload"])
        if payload.get("title") == "Imported property":
            skipped += 1
            continue
        payload["dedupe_fingerprint"] = payload.get("dedupe_fingerprint") or dedupe_fingerprint(payload)
        source_key = (record["source"], record["source_uid"])
        if source_key in seen_source_uids or payload["dedupe_fingerprint"] in seen_fingerprints:
            skipped += 1
            continue
        payload["asset_code"] = f"LJV-{next_code:05d}"
        try:
            prepared = prepare_asset_payload(
                payload,
                source=record["source"],
                source_uid=record["source_uid"],
                source_name=record["source_name"],
                skip_geocode=True,
            )
            asset_row = {
                key: (prepared.get(key) if prepared.get(key) not in ("", []) else None)
                for key in sorted(asset_columns)
            }
            asset_row["approval_status"] = "approved"
            asset_rows.append(asset_row)
            for document in prepared.get("documents", []) or []:
                if isinstance(document, dict):
                    document_rows_by_code.append((payload["asset_code"], document))
            if prepared.get("bottleneck_notes"):
                update_rows_by_code.append((payload["asset_code"], str(prepared["bottleneck_notes"])))
            seen_source_uids.add(source_key)
            seen_fingerprints.add(payload["dedupe_fingerprint"])
            source_counts[record["source"]] += 1
            next_code += 1
        except Exception as exc:
            failed += 1
            failures.append({"source": record["source"], "title": str(payload.get("title") or ""), "error": str(exc)})

    if asset_rows:
        db.execute(models.Asset.__table__.insert(), asset_rows)
        db.commit()
    code_to_id = dict(db.execute(select(models.Asset.asset_code, models.Asset.id)).all())
    document_rows = []
    for asset_code, document in document_rows_by_code:
        asset_id = code_to_id.get(asset_code)
        if not asset_id:
            continue
        document_rows.append(
            {
                "asset_id": asset_id,
                "document_name": document.get("document_name") or document.get("name") or document.get("url") or "Imported collateral",
                "document_type": document.get("document_type") or document.get("type"),
                "url": document.get("url"),
                "storage_path": document.get("storage_path"),
                "notes": document.get("notes"),
            }
        )
    if document_rows:
        db.execute(models.AssetDocument.__table__.insert(), document_rows)
    update_rows = [
        {"asset_id": code_to_id[asset_code], "update_type": "imported_note", "update_text": update_text, "created_by": "quality_reingest"}
        for asset_code, update_text in update_rows_by_code
        if asset_code in code_to_id
    ]
    if update_rows:
        db.execute(models.AssetUpdate.__table__.insert(), update_rows)
    db.execute(
        text(
            """
            INSERT INTO owners (name)
            SELECT DISTINCT owner_name
            FROM (
                SELECT left(NULLIF(trim(COALESCE(raw_source #>> '{source_payload,owner_name}', raw_source->>'OWNER', raw_source->>'OWNER ', raw_source->>'Owner', raw_source->>'Seller')), ''), 255) AS owner_name
                FROM assets
            ) names
            WHERE owner_name IS NOT NULL
              AND NOT EXISTS (SELECT 1 FROM owners o WHERE lower(o.name) = lower(owner_name))
            """
        )
    )
    db.execute(
        text(
            """
            INSERT INTO brokers (name)
            SELECT DISTINCT broker_name
            FROM (
                SELECT left(NULLIF(trim(COALESCE(raw_source #>> '{source_payload,broker_name}', raw_source->>'BROKER', raw_source->>'Broker', raw_source->>'Referrer')), ''), 255) AS broker_name
                FROM assets
            ) names
            WHERE broker_name IS NOT NULL
              AND NOT EXISTS (SELECT 1 FROM brokers b WHERE lower(b.name) = lower(broker_name))
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE assets a
            SET owner_id = o.id
            FROM owners o
            WHERE lower(o.name) = lower(left(NULLIF(trim(COALESCE(a.raw_source #>> '{source_payload,owner_name}', a.raw_source->>'OWNER', a.raw_source->>'OWNER ', a.raw_source->>'Owner', a.raw_source->>'Seller')), ''), 255))
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE assets a
            SET broker_id = b.id
            FROM brokers b
            WHERE lower(b.name) = lower(left(NULLIF(trim(COALESCE(a.raw_source #>> '{source_payload,broker_name}', a.raw_source->>'BROKER', a.raw_source->>'Broker', a.raw_source->>'Referrer')), ''), 255))
            """
        )
    )
    db.commit()
    print(f"bulk_publish_complete created={len(asset_rows)} skipped={skipped}", flush=True)
    return {
        "records_seen": len(records),
        "created": len(asset_rows),
        "skipped_duplicates": skipped,
        "failed": failed,
        "source_counts": dict(source_counts),
        "failures": failures[:20],
    }


def collect_records() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    records.extend(google_sheet_records())
    for workbook in SOURCE_WORKBOOKS:
        if workbook.name == "Pearl Spytech Notes Tracker.xlsx":
            records.extend(pearl_workbook_records(workbook))
        else:
            records.extend(generic_excel_records(workbook))
    return records


def main() -> None:
    create_all()
    with SessionLocal() as db:
        existing_assets = db.scalar(select(func.count(models.Asset.id))) or 0
        if existing_assets:
            backup_path = backup_database(db)
            print(f"backup={backup_path}", flush=True)
        clear_ingested_data(db)
        records = collect_records()
        publish_summary = publish_records(db, records)
        map_summary = enrich_maps(db)
        db.add(
            models.IngestionLog(
                source="quality_reingest",
                filename="google_sheets + uploaded workbooks",
                status="completed",
                total_rows=publish_summary["records_seen"],
                created_count=publish_summary["created"],
                skipped_count=publish_summary["skipped_duplicates"],
            )
        )
        db.commit()
        print(json.dumps({"publish": publish_summary, "maps": map_summary}, indent=2, default=str))


if __name__ == "__main__":
    main()
