from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from backend.app import models
from backend.app.services.asset_ingestor import ingest_or_queue_payload
from backend.app.services.ingestion import load_dedupe_context


COLUMN_ALIASES = {
    "title": ["title", "project", "property", "property name", "asset", "asset name", "deal", "name"],
    "asset_type": ["asset type", "type", "category", "deal type", "purpose"],
    "status": ["status", "stage", "deal status"],
    "source": ["source", "lead source"],
    "locality": ["locality", "location", "micro market", "colony"],
    "area_name": ["area", "area name", "zone"],
    "tehsil": ["tehsil", "taluka"],
    "district": ["district", "city"],
    "state": ["state"],
    "address": ["address", "site address", "full address"],
    "land_area": ["land area", "plot area", "area sq ft", "area acres", "area bigha", "size (for calculation)", "acreage"],
    "built_up_area": ["built up area", "built-up area", "builtup", "constructed area"],
    "asking_price": ["asking price", "ask", "price", "quote", "quoted price"],
    "expected_price": ["expected price", "target price", "expected"],
    "workability_rating": ["workability", "workability rating", "score"],
    "bottleneck_rating": ["bottleneck", "bottleneck rating", "risk rating"],
    "bottleneck_notes": ["bottleneck notes", "risk", "issues", "remarks", "notes", "last update", "reference", "hangup"],
    "legal_status": ["legal", "legal status", "title status"],
    "zoning_status": ["zoning", "zoning status", "land use"],
}

ASSET_TYPE_NORMALIZATION = {
    "land parcel": "land",
    "plot": "land",
    "joint venture": "jv",
    "resale": "resale_unit",
    "resale unit": "resale_unit",
    "commercial": "commercial",
    "rental": "rental",
    "brokerage": "brokerage_listing",
    "brokerage listing": "brokerage_listing",
}


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _normalize_columns(columns: list[str]) -> dict[str, str]:
    lookup = {str(col).strip().lower(): str(col) for col in columns}
    mapping: dict[str, str] = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in lookup:
                mapping[target] = lookup[alias]
                break
    return mapping


def _normalize_asset_type(value: Any) -> str:
    if not value:
        return "land"
    cleaned = str(value).strip().lower()
    if "jv" in cleaned or "joint venture" in cleaned:
        return "jv"
    if "resale" in cleaned:
        return "resale_unit"
    if "rent" in cleaned:
        return "rental"
    if "commercial" in cleaned:
        return "commercial"
    if "plot" in cleaned or "farm" in cleaned or "project" in cleaned or "land" in cleaned:
        return "land"
    return ASSET_TYPE_NORMALIZATION.get(cleaned, cleaned.replace(" ", "_"))


def _is_blank_row(raw: dict[str, Any]) -> bool:
    meaningful = []
    for value in raw.values():
        if value in (None, ""):
            continue
        if isinstance(value, (int, float)) and value == 0:
            continue
        meaningful.append(value)
    return not meaningful


def _parse_coordinates(raw: dict[str, Any]) -> tuple[float | None, float | None]:
    for key, value in raw.items():
        key_lower = str(key).strip().lower()
        if "coordinate" not in key_lower or not value:
            continue
        text = str(value).replace(" ", "")
        if "," not in text:
            continue
        left, right = text.split(",", 1)
        try:
            return float(left), float(right)
        except ValueError:
            continue
    return None, None


def _row_uid(filename: str, row_number: int, raw: dict[str, Any]) -> str:
    digest = hashlib.sha256(f"{filename}:{row_number}:{raw}".encode("utf-8")).hexdigest()[:24]
    return f"excel:{digest}"


def import_excel_to_queue(db: Session, file_path: Path, filename: str) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        header_row = []
        rows = iter(())
    columns = [
        str(value).strip() if value not in (None, "") else f"Column {index + 1}"
        for index, value in enumerate(header_row)
    ]
    mapping = _normalize_columns(columns)
    total = max(sheet.max_row - 1, 0)
    queued = skipped = incomplete = 0

    log = models.IngestionLog(source="excel", filename=filename, status="running", total_rows=total)
    db.add(log)
    db.flush()
    dedupe_context = load_dedupe_context(db)

    for row_number, row in enumerate(rows, start=2):
        raw = {
            columns[index] if index < len(columns) else f"Column {index + 1}": _clean_cell(value)
            for index, value in enumerate(row)
        }
        if _is_blank_row(raw):
            skipped += 1
            continue
        payload: dict[str, Any] = {}
        for target, source_col in mapping.items():
            payload[target] = _clean_cell(raw.get(source_col))
        if not payload.get("title"):
            payload["title"] = payload.get("locality") or payload.get("address")
        latitude, longitude = _parse_coordinates(raw)
        if latitude is not None and longitude is not None:
            payload["latitude"] = latitude
            payload["longitude"] = longitude
            payload["google_maps_link"] = f"https://www.google.com/maps?q={latitude},{longitude}"
        if raw.get("OWNER "):
            payload["owner_name"] = raw.get("OWNER ")
        if raw.get("OWNER"):
            payload["owner_name"] = raw.get("OWNER")
        if raw.get("BROKER"):
            payload["broker_name"] = raw.get("BROKER")
        if raw.get("Area Unit") and payload.get("land_area"):
            payload["land_area"] = f"{payload['land_area']} {raw['Area Unit']}"
        if raw.get("LAST UPDATE") and raw.get("REFERENCE"):
            payload["bottleneck_notes"] = f"{raw['LAST UPDATE']}\n\nHistory: {raw['REFERENCE']}"
        payload["asset_type"] = _normalize_asset_type(payload.get("asset_type"))
        payload["source"] = payload.get("source") or "excel"
        payload["approval_status"] = "pending"
        payload["raw_source"] = raw

        title_parts = [payload.get("title"), payload.get("locality"), payload.get("district")]
        payload["title"] = payload.get("title") or " - ".join(str(part) for part in title_parts if part) or None
        is_incomplete = not payload.get("title") or not (payload.get("address") or payload.get("locality"))
        if is_incomplete:
            incomplete += 1
            payload["needs_manual_review"] = True
            payload["review_reason"] = "Missing title and/or location fields"

        source_uid = _row_uid(filename, int(row_number), raw)
        action, _asset = ingest_or_queue_payload(
            db,
            source="excel",
            source_uid=source_uid,
            title=payload.get("title") or f"Excel row {row_number + 1}",
            payload=payload,
            created_by_source="excel_import",
            dedupe_context=dedupe_context,
        )
        if action in {"created", "queued"}:
            queued += 1
        else:
            skipped += 1

    log.status = "completed"
    log.created_count = queued
    log.review_count = queued
    log.skipped_count = skipped
    db.commit()
    return {
        "source": "excel",
        "total_rows": total,
        "queued_count": queued,
        "skipped_count": skipped,
        "incomplete_count": incomplete,
        "log_id": log.id,
    }
